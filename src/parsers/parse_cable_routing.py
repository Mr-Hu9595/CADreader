#!/usr/bin/env python3
"""线缆路由解析脚本

解析光纤铺设图和桥架铺设图（单次扫描）
提取：光缆型号、路由路径、熔接点、桥架规格

使用方法:
    python parse_cable_routing.py
"""
import csv
import logging
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def parse_all_cable_info(xlsx_path: str, output_dir: Path) -> dict:
    """单次扫描解析所有线缆信息

    Args:
        xlsx_path: Excel文件路径
        output_dir: 输出目录

    Returns:
        统计信息字典
    """
    logger.info(f"解析线缆铺设图: {xlsx_path}")

    # 打开Excel
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # 找到列索引
    text_col_idx = None
    pos_col_idx = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        for j, col_name in enumerate(row):
            if col_name == 'Text String':
                text_col_idx = j
            if col_name == 'Text Position':
                pos_col_idx = j
        break

    if text_col_idx is None:
        logger.error("未找到Text String列")
        wb.close()
        return {'fiber': 0, 'bridge': 0, 'om_os': 0}

    logger.info(f"Text String列索引: {text_col_idx}")

    # 关键词定义
    fiber_keywords = ['光纤', '光缆', '熔接', '配线', 'G.652', 'G.655', 'G.657',
                      '单模', '多模', '尾纤', '跳纤', '收发器', '光模块', '光端机',
                      '12芯', '24芯', '48芯', '96芯']

    bridge_keywords = ['桥架', '槽盒', '托盘', '梯架', '防火', '防腐', '镀锌', '喷塑',
                       '利旧', '新增', '新建', '弯通', '三通', '四通']

    fiber_routes = []
    bridge_routes = []
    om_os_routes = []

    count = 0
    for row in ws.iter_rows(values_only=True):
        count += 1
        if count == 1:  # 跳过表头
            continue
        if count % 100000 == 0:
            logger.info(f"已扫描 {count} 行...")

        cell_value = row[text_col_idx]
        if not cell_value:
            continue

        text = str(cell_value).strip()
        if not text or len(text) < 2:
            continue

        pos = row[pos_col_idx] if pos_col_idx and row[pos_col_idx] else ''

        # 检查光纤关键词
        is_fiber = False
        for kw in fiber_keywords:
            if kw.lower() in text.lower():
                fiber_routes.append({'描述': text, '关键词': kw, '位置': str(pos), '类型': '光纤'})
                is_fiber = True
                break

        # 检查桥架关键词
        if not is_fiber:  # 避免重复
            for kw in bridge_keywords:
                if kw.lower() in text.lower():
                    bridge_routes.append({'描述': text, '关键词': kw, '位置': str(pos), '类型': '桥架'})
                    break

        # 检查OM/OS标识（独立的if，避免与bridge冲突）
        if text.startswith('OM') or text.startswith('OS'):
            # 避免重复添加
            if not any(r['描述'] == text for r in fiber_routes):
                om_os_routes.append({'描述': text, '关键词': text[:2], '位置': str(pos), '类型': '光纤'})

    wb.close()
    logger.info(f"共扫描 {count} 行")

    # 去重
    def dedup(routes):
        seen = set()
        result = []
        for r in routes:
            key = r['描述']
            if key not in seen:
                seen.add(key)
                result.append(r)
        return result

    fiber_routes = dedup(fiber_routes)
    bridge_routes = dedup(bridge_routes)
    om_os_routes = dedup(om_os_routes)

    # 合并光纤和OM/OS（按描述去重）
    all_fiber = dedup(fiber_routes + om_os_routes)

    logger.info(f"光纤相关: {len(all_fiber)} 条")
    logger.info(f"桥架相关: {len(bridge_routes)} 条")

    # 保存结果
    output_dir.mkdir(parents=True, exist_ok=True)

    fiber_path = output_dir / "光纤路由.csv"
    with open(fiber_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['描述', '关键词', '位置', '类型'])
        writer.writeheader()
        writer.writerows(all_fiber)
    logger.info(f"光纤路由已保存: {fiber_path}")

    bridge_path = output_dir / "桥架路由.csv"
    with open(bridge_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['描述', '关键词', '位置', '类型'])
        writer.writeheader()
        writer.writerows(bridge_routes)
    logger.info(f"桥架路由已保存: {bridge_path}")

    return {'fiber': len(all_fiber), 'bridge': len(bridge_routes), 'om_os': len(om_os_routes)}


def main():
    """主函数"""
    base_dir = Path(r"D:\工作\日常工作\首山碳材环保管控平台")
    output_dir = base_dir / r"data\图纸解析结果\网络拓扑"

    # 光纤/网络铺设图
    fiber_xlsx = base_dir / r"首山四期图纸\首山焦化四期初步设计图纸-20251209\2、网络设计\首山碳材料治理项目光纤铺设图251031_dwg.xlsx"

    if not fiber_xlsx.exists():
        logger.error(f"文件不存在: {fiber_xlsx}")
        print(f"错误: 光纤铺设图不存在")
        return

    # 执行解析
    stats = parse_all_cable_info(str(fiber_xlsx), output_dir)

    print(f"\n=== 解析结果 ===")
    print(f"光纤相关标注: {stats['fiber']} 条")
    print(f"桥架相关标注: {stats['bridge']} 条")
    print(f"OM/OS标识: {stats['om_os']} 条")
    print(f"\n输出目录: {output_dir}")


if __name__ == '__main__':
    main()
